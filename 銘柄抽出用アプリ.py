from yahoo_finance_api2 import share
from yahoo_finance_api2.exceptions import YahooFinanceError
from datetime import date,datetime,timedelta
import pandas as pd 
import numpy as np
import openpyxl as xl
import statistics as st

# 日経225の銘柄コードをテキストから取得
f = open('日経225.txt','r', encoding='UTF-8')
nikkeiCodeList = f.readlines()

UpCode1 = [] #現物買い　1：5日上昇かつ5日>20日かつ60日下落
UpCode2 = [] #現物買い　2：5日下落かつ5日<20日かつ60日下落
UpCode3 = [] #現物買い　3：5日下落かつ5日>20日かつ60日下落
DownCode1 = [] #空売り　 4：5日下落かつ5日<20日かつ60日水平
DownCode2 = [] #空売り　 5：5日下落かつ5日<20日かつ60日下落

# 株価データを取得
for list in range(len(nikkeiCodeList)):
  #テキストを一行ずつ読み取ったとき/nという文字列がはいるため削除する
  nikkeiCode = nikkeiCodeList[list].rstrip('\n')
  nikkeiCodeT = share.Share(nikkeiCode + ".T")

  try:
    nikkeiData = nikkeiCodeT.get_historical(share.PERIOD_TYPE_MONTH,5,share.FREQUENCY_TYPE_DAY,1)
  except YahooFinanceError as e:
    print(e.message)

  #print(nikkeiData)

  nikkeiData = pd.DataFrame({'datetime': [datetime.fromtimestamp(d / 1000).date() for d in nikkeiData['timestamp']],\
    'open' : nikkeiData['open'], 'high' : nikkeiData['high'],\
    'close' : nikkeiData['close'], 'volume' : nikkeiData['volume']})
  nikkeiData = nikkeiData.sort_values('datetime', ascending= False)

  #終わり値の抽出
  endPrice = nikkeiData['close'].values
  #print(nikkeiData['datetime'].head(1).values)
  #5日移動平均線の値を計算
  SMA5 = np.convolve(endPrice,np.ones(5), mode='valid') / 5
  #リスト項目数に合わせるため、計算から漏れる値を補完
  SMA5 = np.append(SMA5,[0,0,0,0])
  #nikkeiDataの2次配列に5日移動平均線の値を追加
  nikkeiData['SMA5'] = SMA5

  #20日移動平均線の計算式
  SMA20 = np.convolve(endPrice,np.ones(20), mode='valid') / 20
  #リスト項目数に合わせるため、計算から漏れる値を補完
  REPEAT_COUNT = 19
  l = [0]
  for n in range(REPEAT_COUNT):
    for elem in l:
      SMA20 = np.append(SMA20,elem)
  #nikkeiDataの2次配列に5日移動平均線の値を追加
  nikkeiData['SMA20'] = SMA20

  #60日移動平均線の計算式
  SMA60 = np.convolve(endPrice,np.ones(60), mode='valid') / 60
  #リスト項目数に合わせるため、計算から漏れる値を補完
  REPEAT_COUNT = 59
  l = [0]
  for n in range(REPEAT_COUNT):
    for elem in l:
      SMA60 = np.append(SMA60,elem)
  #nikkeiDataの2次配列に5日移動平均線の値を追加
  nikkeiData['SMA60'] = SMA60
  #print(nikkeiData)


  """
  スクリーニングルール
  出来高株数中央値(statistics.median(変数))>500,000

    (現物買い)
    SAM5>=SAM20かつSAM5<SAM60のとき
      1 SAM60['78']>SAM60['99'] 下落のとき (SMA60['78']=1ヶ月前)
        close['96']+(close['96']*0.05)>close['99'] かつ　close['96']-(close['96']*0.05)<close['99'] (終値が3日前の終わり値の5%未満の変化のとき)
        
    
    SAM5<=SAM20かつSAM5>SAM60のとき
      2 SAM60['78']+(SAM60['78']*0.01)>SAM60['99']かつSAM60['78']-(SAM60['78']*0.01)>SAM60['99'] 
       (60日線下落のとき SAM['78']=約1ヶ月前、株価の前後1%のズレを許容)

      3 SAM60['78']+(SAM60['78']*0.01)<SAM60['99']かつSAM60['78']-(SAM60['78']*0.01)<SAM60['99']
       (60日線上昇のとき)

    (空売り)
    SAM5<=SAM20かつSAM5>SAM60のとき
      4 SAM5['96']>SAM20['96']またはSAM5['97']>SAM20['97']またはSAM5['98']>SAM20['98']
       （3日前から5日線が20日線を割り始めたもの）
    
    SAM5<SAM20かつSAM5<SAM60のとき
      5 SAM60['78']+(SAM60['78']*0.01)>SAM60['99']かつSAM60['78']-(SAM60['78']*0.01)>SAM60['99'] 
       (60日線下落のとき SAM['78']=約1ヶ月前、株価の前後1%のズレを許容)

  """
  medianVolume = st.median(nikkeiData['volume'])
  #print(pd.to_numeric(nikkeiData['datetime']).idxmax())
  #print(SMA5Max)

  #データカラム数の取得と2~4日前の取得、前月の取得
  lenMax = len(nikkeiData['open']) - 1 
  len2Be = lenMax - 2
  len3Be = lenMax - 3
  len4Be = lenMax - 4
  len20Be = lenMax - 20

  #前日の各平均線の値
  SMA5b1 = nikkeiData['SMA5'][lenMax]
  SMA20b1 = nikkeiData['SMA20'][lenMax]
  SMA60b1 = nikkeiData['SMA60'][lenMax]

  #1ヶ月前の60日線の金額
  SMA60b20 = nikkeiData['SMA60'][len20Be]

  #2~4日前の5日線、20日線の金額
  SMA5b2 = nikkeiData['SMA5'][len2Be]
  SMA5b3 = nikkeiData['SMA5'][len3Be]
  SMA5b4 = nikkeiData['SMA5'][len4Be]
  SMA20b2 = nikkeiData['SMA20'][len2Be]
  SMA20b3 = nikkeiData['SMA20'][len3Be]
  SMA20b4 = nikkeiData['SMA20'][len4Be]

  #前日終値と3日前の終値
  closeb1 = nikkeiData['close'][lenMax]
  closeb4 = nikkeiData['close'][len4Be]

  if medianVolume > 500000:
    if closeb1 < 4000:
        if(SMA5b1 >= SMA20b1) and (SMA5b1 < SMA60b1):#SAM5>=SAM20かつSAM5<SAM60のとき
            if SMA60b20 > SMA60b1:
                if (closeb4+(closeb4*0.03) > closeb1) and (closeb4-(closeb4*0.03) < closeb1):#ボックス相場
                    UpCode1.append([nikkeiCode,closeb1])
                    #1：5日上昇かつ5日>20日かつ60日下落（現物買い）

        if(SMA5b1 <= SMA20b1) and (SMA5b1 > SMA60b1):#SAM5<=SAM20かつSAM5>SAM60のとき
            if (SMA60b20-(SMA60b20*0.01) > SMA60b1):
                UpCode2.append([nikkeiCode,closeb1])
                #2：5日下落かつ5日<20日かつ60日下落（現物買い）

            if (SMA60b20+(SMA60b20*0.01) < SMA60b1):
                UpCode3.append([nikkeiCode,closeb1])
                #3：5日下落かつ5日>20日かつ60日上昇（現物買い）

            if(SMA5b3 > SMA20b3) or (SMA5b2 > SMA20b2):
                if(SMA60b20+(SMA60b20*0.01) > SMA60b1) and (SMA60b20-(SMA60b20*0.01) < SMA60b1):
                    DownCode1.append([nikkeiCode,closeb1])
                    #4：5日下落かつ5日<20日かつ60日水平

        if(SMA5b1 < SMA20b1) and (SMA5b1 < SMA60b1):#SAM5<SAM20かつSAM5<SAM60のとき
            if (SMA60b20-(SMA60b20*0.01) > SMA60b1):
                DownCode2.append([nikkeiCode,closeb1])
                #5：5日下落かつ5日<20日かつ60日下落

wb = xl.load_workbook('(自動)銘柄抽出結果.xlsx')

#前回のNEWシートを昨日日付、グレーに変更
ws = wb['TODAY']
yesterday = date.today() - timedelta(days=1) 
ws.title = f'{yesterday:%Y%m%d}'
ws.sheet_properties.tabColor = '808080'#gray
wb.move_sheet(ws, offset=wb.index(wb['テンプレ']))

#テンプレシートをコピーしてNEWシートの複製
copyWs = wb.copy_worksheet(wb['テンプレ'])
copyWs.title = 'TODAY'
copyWs.sheet_properties.tabColor = 'FFA500'#orange
wb.move_sheet(copyWs, offset=-(len(wb.worksheets)-1))#先頭まで移動

#NEWのシートを取得して、先頭にあるシートを選択、その後、更新
wsNew = wb['TODAY']
wb.active = 0

if len(UpCode1) != 0:
  for y, row in enumerate(UpCode1):
    for x, cell in enumerate(row):
        wsNew.cell(row= y+3, column= x+1, value=UpCode1[y][x])

if len(UpCode2) != 0:
  for y, row in enumerate(UpCode2):
    for x, cell in enumerate(row):
        wsNew.cell(row= y+3, column= x+3, value=UpCode2[y][x])

if len(UpCode3) != 0:
  for y, row in enumerate(UpCode3):
    for x, cell in enumerate(row):
        wsNew.cell(row= y+3, column= x+5, value=UpCode3[y][x])

if len(DownCode1) != 0:
  for y, row in enumerate(DownCode1):
    for x, cell in enumerate(row):
        wsNew.cell(row= y+3, column= x+7, value=DownCode1[y][x])

if len(DownCode2) != 0:
  for y, row in enumerate(DownCode2):
    for x, cell in enumerate(row):
        wsNew.cell(row= y+3, column= x+9, value=DownCode2[y][x])

wb.save('(自動)銘柄抽出結果.xlsx')
wb.close()
